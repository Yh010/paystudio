from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from typing import Optional

app = FastAPI()

def process_excel(input_file: str, output_file: str):
    # Read the Excel file based on extension
    if input_file.endswith('.xlsx'):
        # For .xls files, we need to read the specific range
        wb = load_workbook(input_file, data_only=True)
        ws = wb.active
        
        # Get headers from B13:I13
        headers = []
        for cell in ws['B13:I13'][0]:
            headers.append(cell.value)
        
        # Get data starting from row 14
        data = []
        for row in ws.iter_rows(min_row=14, min_col=2, max_col=9):
            row_data = [cell.value for cell in row]
            if any(row_data):  # Only add non-empty rows
                data.append(row_data)
        
        df = pd.DataFrame(data, columns=headers)
    else:
        df = pd.read_excel(input_file, engine='openpyxl')
    
    # Print column names for debugging
    print("Available columns:", df.columns.tolist())
    
    # Find the transaction remarks column (case-insensitive)
    remarks_col = None
    for col in df.columns:
        if 'transaction' in str(col).lower() and 'remark' in str(col).lower():
            remarks_col = col
            break
    
    if remarks_col is None:
        raise ValueError("Could not find Transaction Remarks column in the Excel file")
    
    # Process Transaction Remarks column
    def extract_third_value(remark: str) -> str:
        if pd.isna(remark):
            return ""
        parts = str(remark).split('/')
        return parts[2] if len(parts) > 2 else remark
    
    # Apply the processing
    df[remarks_col] = df[remarks_col].apply(extract_third_value)
    
    # Save to new Excel file
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    # Apply yellow highlighting for non-zero deposit amounts
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Find the Deposit Amount column (case-insensitive)
    deposit_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value and 'deposit' in str(cell.value).lower() and 'amount' in str(cell.value).lower():
            deposit_col = idx
            break
    
    if deposit_col:
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Start from row 2 (skip header)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=deposit_col)
            if cell.value and float(cell.value) != 0.0:
                # Fill entire row with yellow
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = yellow_fill
    
    wb.save(output_file)

@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    # Save uploaded file
    input_path = f"temp_input{os.path.splitext(file.filename)[1]}"
    output_path = "processed_output.xlsx"
    
    with open(input_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    try:
        # Process the file
        process_excel(input_path, output_path)
        
        # Return processed file
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="processed_statement.xlsx"
        )
    finally:
        # Clean up input file
        if os.path.exists(input_path):
            os.remove(input_path)

@app.get("/")
async def read_root():
    return {"message": "Welcome to Bank Statement Processor API"} 